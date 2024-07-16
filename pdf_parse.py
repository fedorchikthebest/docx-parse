from docx_parser import DocumentParser
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

from string import ascii_uppercase
from utils import coords_to_xl, to_name
import sys


def pdf_parse(name):
    doc = DocumentParser(f'./upload/{name}')
    a = 0
    OTSTUP = 2

    SIZE_ROW = 86.25

    text_font = Font(name='Courier New',
                    size=10,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,)

    head_font = Font(name='Courier New',
                    size=10,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,)

    text_aligment = Alignment(
        horizontal='left',
        vertical='top',
        wrapText=True
    )

    head_aligment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True
    )

    thin = Side(border_style="thin", color="000000")

    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    data = []
    formated_data = []
    sizes = [17.14, 14, 17.14, 11.86, 18, 21, 10, 24.14]

    for _type, item in doc.parse():
        if _type == "table":
            if a == 1:
                data = item.get("data")
                break
            a += 1

    wb = Workbook()
    ws = wb.active

    ws.title = 'zapros'

    ws['A1'] = 'Приложение к запросу'

    ws['A2'] = 'Фамилия'
    ws['B2'] = 'Имя'
    ws['C2'] = 'Отчество'
    ws['D2'] = 'Дата Рождения'
    ws['E2'] = 'Место рождения'
    ws['F2'] = 'Адрес регистрации и (или) фактического места проживания'
    ws['G2'] = 'Серия и номер паспорта'
    ws['H2'] = 'Дата выдачи и наименование органа, выдавшего паспорт'

    for i in range(8):
        ws.column_dimensions[ascii_uppercase[i]].width = sizes[i]
        ws[coords_to_xl(i, 1)].font = head_font
        ws[coords_to_xl(i, 1)].border = border
        ws[coords_to_xl(i, 1)].alignment = head_aligment
        
    for y, val_y in enumerate(filter(lambda x: x[0] == '', data)):
        if not val_y[0]:
            p_ser, p_num = val_y[5].split(', ')[0].split()
            vidacha = ', '.join(val_y[5].split(', ')[1:])
        
        formated_data.append(to_name(val_y[1].split()) + val_y[2:5] + [f'{p_ser} № {p_num}', vidacha])

    for y, val_y in enumerate(formated_data):
        for x, val_x in enumerate(val_y):
            ws[coords_to_xl(x, y + OTSTUP)] = val_x
            ws[coords_to_xl(x, y + OTSTUP)].font = text_font
            ws[coords_to_xl(x, y + OTSTUP)].border = border
            ws[coords_to_xl(x, y + OTSTUP)].alignment = text_aligment

    wb.save(f'./static/{".".join([name.split(".")[0], "xlsx"])}')
    return '.'.join([name.split('.')[0], 'xlsx'])